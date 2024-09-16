import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
	filename = 'Summary Of Worker Total Salary Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	title_1=title1(args)
	ws.append(title_1)
	bold_font = Font(bold=True)
	for cell in ws["1:1"]:
		cell.font = bold_font
	ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=25)
	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=25):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	title_2=title2(args)
	ws.append(title_2)
	bold_font = Font(bold=True)
	for cell in ws["2:2"]:
		cell.font = bold_font
	ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=25)
	for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=25):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	header_1=header1(args)
	ws.append(header_1)
	bold_font = Font(bold=True)
	for cell in ws["3:3"]:
		cell.font = bold_font
	
	for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=25):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	
	bold_font = Font(bold=True)
	data_1 = data1(args)
	for data in data_1:
		ws.append(data)
	ws.merge_cells(start_row=5,start_column=1,end_row=5,end_column=2)

	bold_font = Font(bold=True)
    
	for cell in ws["5:5"]:
		cell.font = bold_font
	for rows in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=2):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	border = Border(left=Side(border_style='thin', color='000000'),
			right=Side(border_style='thin', color='000000'),
			top=Side(border_style='thin', color='000000'),
			bottom=Side(border_style='thin', color='000000')) 
	for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+3, min_col=1, max_col=25):
		for cell in rows:
			cell.border = border
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
	data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
	return data
@frappe.whitelist()
def title2(args):
	start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
	mon = start_date.strftime('%B %Y').upper()
	data = ['SUMMARY OF Worker SALARY FOR THE MONTH OF %s' % mon]
	return data
@frappe.whitelist()
def header1(args):
	data = ['Sr.No.','Departments','No.Of Employees','Basic Salary','Dearness Allowance','Arr.Basic Salary & DA','Special Allowance','Incentives','Arrears Salary','Over Time Salary','Arr. Over Time Salary','Shift Allowance','Gross Salary','PF','PT','ESI','LWF','Salary Deduction','Deductions','Attd. Bonus','Casual Leave Encashment','Earned Leave Encashment','Extra Pay','Bonus Yearly','Net Salary']
	return data
@frappe.whitelist()
def data1(args):
	data = []
	sa = frappe.get_all('Salary Slip', {'salary_structure': "Worker",'start_date':args.start_date,'end_date':args.end_date}, ['*'])
	total_employee_count = len(sa)  
	for i in sa:
		basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].basic or 0
		da = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as da from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Dearness Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].da or 0
		arb = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as arb from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Arrear Basic' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].arb or 0
		sa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].sa or 0
		i = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as i from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Incentives' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].i or 0
		a= frappe.db.sql("""select sum(`tabSalary Detail`.amount) as a from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].a or 0
		ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Overtime' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].ot or 0
		aot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aot from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Arrear Overtime' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].aot or 0
		sha = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sha from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].sha or 0
		gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
		where `tabSalary Slip`.salary_structure = 'Worker'and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.from_date,args.to_date),as_dict=True)[0].gross or 0
		pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Provident Fund' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].pf or 0
		pt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pt from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Professional Tax' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].pt or 0
		esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].esi
		lwf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lwf from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'LWF' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].lwf or 0
		sd = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sd from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Salary Deduction' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].sd or 0
		total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
		where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].total_deduct or 0		
		ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].ab or 0
		cle = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as cle from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Casual Leave Encashment' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].cle or 0
		ele = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ele from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Earned Leave Encashment' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].ele  or 0
		ep = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ep from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Extra Pay' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].ep or 0
		b = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as b from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.salary_structure = 'Worker' and `tabSalary Detail`.salary_component = 'Bonus' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].b or 0
		npay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as npay from `tabSalary Slip`
		where `tabSalary Slip`.salary_structure = 'Worker'and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].npay or 0
		row = ['1', 'Worker', total_employee_count, basic, da, arb, sa, i, a, ot, aot, sha, gross, pf, pt, esi, lwf, sd, total_deduct, ab, cle, ele, ep, b, npay]

	data.append(row)
	data.append(["Grand Total","",total_employee_count, basic, da, arb, sa, i, a, ot, aot, sha, gross, pf, pt, esi, lwf, sd, total_deduct, ab, cle, ele, ep, b, npay])
	return data