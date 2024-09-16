import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Staff Salary - Allocation to Cost Center Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['M'].width = 35
    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=13)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    title_2=title2(args)
    ws.append(title_2)
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=13)
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=13):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=13):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data_1 = data1(args)
    for data in data_1:
        ws.append(data)
    for rows in ws.iter_rows(min_row=3, max_row=len(data_1)+3, min_col=2, max_col=13):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    for header in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=9):
         for cell in header:
             cell.fill = PatternFill(fgColor='B9DAFF', fill_type = "solid")
    for header in ws.iter_rows(min_row=3, max_row=3, min_col=10, max_col=13):
         for cell in header:
             cell.fill = PatternFill(fgColor='AECC94', fill_type = "solid")
    for header in ws.iter_rows(min_row=len(data_1)+3, max_row=len(data_1)+3, min_col=1, max_col=9):
         for cell in header:
             cell.fill = PatternFill(fgColor='B9DAFF', fill_type = "solid")
    for header in ws.iter_rows(min_row=len(data_1)+3, max_row=len(data_1)+3, min_col=10, max_col=13):
         for cell in header:
             cell.fill = PatternFill(fgColor='AECC94', fill_type = "solid")
    for header in ws.iter_rows(min_row=len(data_1)+3, max_row=len(data_1)+3, min_col=1, max_col=1):
         for cell in header:
             cell.font = Font(bold=True)
    for header in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=1):
         for cell in header:
             cell.font = Font(bold=True)
    for header in ws.iter_rows(min_row=len(data_1)+2, max_row=len(data_1)+2, min_col=11, max_col=11):
         for cell in header:
             cell.font = Font(bold=True)
    for header in ws.iter_rows(min_row=len(data_1), max_row=len(data_1), min_col=1, max_col=11):
         for cell in header:
             cell.font = Font(bold=True)
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+3, min_col=1, max_col=13):
        for cell in rows:
            cell.border = border

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
    data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
    return data
@frappe.whitelist()
def title2(args):
    month = datetime.strptime(str(args.start_date),'%Y-%m-%d')
    mon = str(month.strftime('%b') +' '+ str(month.strftime('%Y')))
    data = ['Staff Salary - Allocation to cost centers']
    return data
@frappe.whitelist()
def header1(args):
    data = ['Departments','Manpower Count','Incentives','Arrears Salary','Over Time Salary','Arr OT Salary','Shift Allowance','Gross Salary','Attendance Bonus','Cost Center','Cost Center Name','Gross Salary','Over Time & Arrear OT and DH Bonus']
    return data

@frappe.whitelist()
def data1(args):
    data = []
    salary_structure_count = {}
    Dep = frappe.get_all('Department', {'parent_department': 'SG & A'}, ['department_name'])
    data.append(['SG & A(a)'])
    for i in Dep:
        dm = i.department_name
        frappe.log_error(dm)

        row=['SG & A',]
        row = [ dm]
        # data.append(row)
    sa = frappe.get_all('Salary Slip', {'start_date': args.start_date, 'end_date': args.end_date, 'salary_structure': 'Worker','department':dm}, ['department'])
    
    department_totals = {}
    salary_structure_count = {}

    department_to_cost_center = {}
    for p in sa:
        ss = p.department
        frappe.log_error(ss)

        structure_escaped = frappe.db.escape(ss)

        cost_center = frappe.get_value('Department', ss, 'payroll_cost_center')
        frappe.log_error(cost_center)

        department_to_cost_center[ss] = cost_center

        def get_salary_component_total(salary_component):
            return frappe.db.sql(f"""
                select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
                left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
                where `tabSalary Slip`.department = {structure_escaped} and
                    `tabSalary Detail`.salary_component = '{salary_component}' and
                    `tabSalary Slip`.start_date >= '{args.start_date}' and
                    `tabSalary Slip`.end_date <= '{args.end_date}'
            """, as_dict=True)[0].total or 0

        i = get_salary_component_total('Incentives')
        a = get_salary_component_total('Arrear')
        ot = get_salary_component_total('Overtime')
        aot = get_salary_component_total('Arrear Overtime')
        sha = get_salary_component_total('Shift Allowance')

        gross = frappe.db.sql(f"""
            select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
            where `tabSalary Slip`.department = {structure_escaped} and
                `tabSalary Slip`.start_date >= '{args.start_date}' and
                `tabSalary Slip`.end_date <= '{args.end_date}'
        """, as_dict=True)[0].gross or 0

        ab = get_salary_component_total('Attendance Bonus')
        gross_amt = gross + ab  # Calculate Gross Salary + Attendance Bonus
        hra = get_salary_component_total('DH Allowance')

        if ss in salary_structure_count:
            salary_structure_count[ss] += 1
        else:
            salary_structure_count[ss] = 1

        if ss in department_totals:
            department_totals[ss][0] += i
            department_totals[ss][1] += a
            department_totals[ss][2] += ot
            department_totals[ss][3] += aot
            department_totals[ss][4] += sha
            department_totals[ss][5] += gross
            department_totals[ss][6] += ab
            department_totals[ss][7] += gross_amt
            department_totals[ss][8] += hra
        else:
            department_totals[ss] = [i, a, ot, aot, sha, gross, ab, gross_amt, hra]

    total_i = sum(x[0] for x in department_totals.values())
    total_a = sum(x[1] for x in department_totals.values())
    total_ot = sum(x[2] for x in department_totals.values())
    total_aot = sum(x[3] for x in department_totals.values())
    total_sha = sum(x[4] for x in department_totals.values())
    total_gross = sum(x[5] for x in department_totals.values())
    total_ab = sum(x[6] for x in department_totals.values())
    total_gross_amt = sum(x[7] for x in department_totals.values())
    total_hra = sum(x[8] for x in department_totals.values())

    for ss, totals in department_totals.items():
        cc = department_to_cost_center[ss]
        vv=[totals[2]+totals[3]+totals[8]]
        data.append([ss, salary_structure_count[ss]] + [totals[0]]+ [totals[1]]+ [totals[2]]+ [totals[3]]+ [totals[4]]+ [totals[5]]+ [totals[6]]+[cc]+[ss]+ [totals[7]]+vv)
    salary_structure_counts = [salary_structure_count[ss] for ss in department_totals.keys()]
    total_counts = sum(salary_structure_counts)
    total_vv=total_ot+total_aot+total_hra
    grand_total_row = ['Indirect(b)','','','','','','','','','','Total(a)'] +[total_gross_amt,total_vv]
    data.append( grand_total_row) 
    Deps = frappe.get_all('Department', {'parent_department': 'Indirect'}, ['department_name'])
    for j in Deps:
        vv = j.department_name
        frappe.errprint(vv)
    sb = frappe.get_all('Salary Slip', {'start_date': args.start_date, 'end_date': args.end_date, 'salary_structure': 'Worker','department':vv}, ['department'])
    
    department_totals = {}
    salary_structure_count = {}

    department_to_cost_center = {}
    for k in sb:
        rr = k.department
        frappe.log_error(rr)
        structure_escaped = frappe.db.escape(rr)

        cost_center = frappe.get_value('Department', rr, 'payroll_cost_center')
        frappe.log_error(cost_center)

        department_to_cost_center[rr] = cost_center

        def get_salary_component_total(salary_component):
            return frappe.db.sql(f"""
                select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
                left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
                where `tabSalary Slip`.department = {structure_escaped} and
                    `tabSalary Detail`.salary_component = '{salary_component}' and
                    `tabSalary Slip`.start_date >= '{args.start_date}' and
                    `tabSalary Slip`.end_date <= '{args.end_date}'
            """, as_dict=True)[0].total or 0

        i = get_salary_component_total('Incentives')
        a = get_salary_component_total('Arrear')
        ot = get_salary_component_total('Overtime')
        aot = get_salary_component_total('Arrear Overtime')
        sha = get_salary_component_total('Shift Allowance')

        gross = frappe.db.sql(f"""
            select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
            where `tabSalary Slip`.department = {structure_escaped} and
                `tabSalary Slip`.start_date >= '{args.start_date}' and
                `tabSalary Slip`.end_date <= '{args.end_date}'
        """, as_dict=True)[0].gross or 0

        ab = get_salary_component_total('Attendance Bonus')
        gross_amt = gross + ab  # Calculate Gross Salary + Attendance Bonus
        hra = get_salary_component_total('DH Allowance')

        if ss in salary_structure_count:
            salary_structure_count[rr] += 1
        else:
            salary_structure_count[rr] = 1

        if rr in department_totals:
            department_totals[rr][0] += i
            department_totals[rr][1] += a
            department_totals[rr][2] += ot
            department_totals[rr][3] += aot
            department_totals[rr][4] += sha
            department_totals[rr][5] += gross
            department_totals[rr][6] += ab
            department_totals[rr][7] += gross_amt
            department_totals[rr][8] += hra
        else:
            department_totals[rr] = [i, a, ot, aot, sha, gross, ab, gross_amt, hra]

    total_i = sum(x[0] for x in department_totals.values())
    total_a = sum(x[1] for x in department_totals.values())
    total_ot = sum(x[2] for x in department_totals.values())
    total_aot = sum(x[3] for x in department_totals.values())
    total_sha = sum(x[4] for x in department_totals.values())
    total_gross = sum(x[5] for x in department_totals.values())
    total_ab = sum(x[6] for x in department_totals.values())
    total_gross_amt1 = sum(x[7] for x in department_totals.values())
    total_hra = sum(x[8] for x in department_totals.values())

    for rr, totals in department_totals.items():
        cc = department_to_cost_center[rr]
        vv=[totals[2]+totals[3]+totals[8]]
        data.append([rr, salary_structure_count[rr]] + [totals[0]]+ [totals[1]]+ [totals[2]]+ [totals[3]]+ [totals[4]]+ [totals[5]]+ [totals[6]]+[cc]+[rr]+ [totals[7]]+vv)
    salary_structure_counts = [salary_structure_count[rr] for rr in department_totals.keys()]
    total_counts = sum(salary_structure_counts)
    total_vv1=total_ot+total_aot+total_hra
    grand_total_row1 = ['','','','','','','','','','','Total(b)'] +[total_gross_amt1,total_vv1]
    data.append( grand_total_row1)
    total=total_gross_amt1+total_gross_amt
    total1=total_vv1+total_vv
    data.append(['Total(a+b)','','','','','','','','','','',total,total1])
    return data