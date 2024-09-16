from __future__ import unicode_literals
import frappe
from calendar import monthrange
from frappe import _, msgprint
from frappe.utils import flt
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download():
	filename = 'PF Report'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['B'].width = 13 
	ws.column_dimensions['C'].width = 20 
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 25
	ws.column_dimensions['F'].width = 10
	ws.column_dimensions['G'].width = 8
	ws.append(["S.No","Employee ID","Employee Name","Department","Bank Account Number","Gross Pay","PF"])
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	data1= get_data(args)
	for row in data1:
		ws.append(row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_data(args):
	data = []
	row = []
	salary_slips = frappe.db.sql("""select * from `tabSalary Slip` where start_date between '%s' and '%s' and docstatus!=2"""%(args.start_date, args.end_date), as_dict=True)
	i = 1
	for ss in salary_slips:
		emp_id = ss.employee
		emp_name = ss.employee_name
		dep= ss.department
		ac = ss.bank_account_no
		sal = ss.gross_pay
		pf = frappe.get_value('Salary Detail',{'salary_component':"Provident Fund",'parent':ss.name},["amount"] or 0)
		row = [i,emp_id or "-",emp_name or "-",dep or "-",ac or "-",sal or 0,pf or 0]
		data.append(row)
		i+=1
	return data