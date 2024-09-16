import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download(start_date,end_date,department=None):
	filename = 'Monthly In/Out Report'
	args = {'start_date':start_date,'end_date':end_date,'department':department}
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)

def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 54
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 15
	ws.column_dimensions['D'].width = 15
	ws.column_dimensions['E'].width = 12
	ws.append(['T S INTERSEATS INDIA PVT LTD'])

	header_date = title1(args)
	ws.append(title1(args))

	header_date = get_dep(args)
	ws.append(get_dep(args))

	header_date = get_day(args)
	ws.append(get_day(args))

	emp= get_data(args)
	for e in emp:
		ws.append(e)


	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.log_error(title='test',message=xlsx_file)
	return xlsx_file 

def build_xlsx_response(filename,args):
	frappe.log_error(title='test',message=filename)
	xlsx_file = make_xlsx(filename,args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Reports Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	frappe.db.set_value('Reports Dashboard',None,'attach',attached_file.file_url)


def title1(args):
	fdate = datetime.strptime(args['start_date'],'%Y-%m-%d')
	tdate = datetime.strptime(args['end_date'],'%Y-%m-%d')
	f_date = fdate.strftime('%d/%m/%Y')
	t_date = tdate.strftime('%d/%m/%Y')
	data = ["Monthly IN OUT Report for the Period of " +f_date + " To " +t_date]
	
	return data

@frappe.whitelist()
def get_dep(args):
	data = []
	data += ['Employee ID','Employee Name','Designation','Department','DOJ','Date']
	dates = get_dates(args)
	for date in dates:
		date = datetime.strptime(date,'%Y-%m-%d')
		date = date.strftime('%d')
		data.extend([date])
	data.extend(["Present","Half Day","On Duty","Permission","Absent","Weekoff","Holiday","Paid Leave","LOP","COFF","OT","Actual Late",'Late Deduct','Permission','Night Shift','Total Working Days'])
	return data

def get_dates(args):
	no_of_days = date_diff(add_days(args['end_date'], 1), args['start_date'])
	dates = [add_days(args['start_date'], i) for i in range(0, no_of_days)]
	return dates

@frappe.whitelist()
def get_day(args):
	data = []
	data += ['','','','','','']
	dates = get_dates(args)
	for date in dates:
		date = datetime.strptime(date,'%Y-%m-%d')
		date = date.strftime('%A')
		data.extend([date])
	return data

def get_data(args):
	data = []
	if 'employee' in args:
		employees = frappe.get_all("Employee",{'name':args['employee'],'status':'Active'},['*'])
	# elif args.employee and args.employee_category:
	# 	employees = frappe.get_all("Employee",{'name':args['employee'],'employee_catagory':args['employee_category'],'status':'Active'},['*'])
	else:
		employees = frappe.get_all("Employee",{'status':'Active'},['*'])
	for emp in employees:
		start_date = args.get('start_date')
		end_date = args.get('end_date')
		dates = get_dates(args)
		doj_crt = emp.date_of_joining.strftime('%d-%m-%Y')
		row1 = [emp.name,emp.employee_name,emp.designation,emp.department,doj_crt,"Status"]
		row2 = ["","","","","","Shift"]
		row3 = ["","","","","","In Time"]
		row4 = ["","","","","","Out Time"]
		row5 = ["","","","","","Late"]
		row6 = ["","","","","","Early"]
		row7 = ["","","","","","TWH"]
		row8 = ["","","","","","OT"]
		total_present = 0
		total_half_day = 0
		total_absent = 0
		total_holiday = 0
		total_weekoff = 0
		total_ot = 0
		total_od = 0
		total_permission = 0
		total_lop = 0
		total_paid_leave = 0
		total_combo_off = 0
		n_shift = 0
		total_late = timedelta(0,0,0)
		total_late_deduct = timedelta(0,0)
		ww = 0
		twh = 0
		ot = 0
		for date in dates:
			if frappe.db.exists("Attendance",{'attendance_date':date,'employee':emp.name,'docstatus':('!=','2')}):
				att = frappe.db.sql("""select * from `tabAttendance` where attendance_date = '%s' and employee = '%s' """%(date,emp.name),as_dict=1)
				if att[0].get('shift_status'):
					row1.append(att[0].get('shift_status'))
					if att[0].get('shift_status') =="P":
						total_present += 1
						# cell.font = Font(bold=True,size=14)
					elif att[0].get('shift_status') == "P/N":
						total_present +=1
						n_shift +=1
					elif att[0].get('shift_status') == "P/AB":
						total_present +=0.5
						total_half_day +=0.5
					elif att[0].get('shift_status') == "DH/P":
						total_present +=1
					elif att[0].get('shift_status') == "DH/PN":
						total_present +=1
						n_shift +=1
					elif att[0].get('shift_status') == "DH/PP":
						total_present +=0.5
					elif att[0].get('shift_status') == "CL":
						total_paid_leave +=1
					elif att[0].get('shift_status') == "P/CL":
						total_paid_leave +=0.5
						total_present +=0.5
					elif att[0].get('shift_status') == "EL":
						total_paid_leave +=1
					elif att[0].get('shift_status') == "P/EL":
						total_paid_leave +=0.5
						total_present +=0.5
					elif att[0].get('shift_status') == "ESI":
						esi +=1
					elif att[0].get('shift_status') == "LOP":
						total_lop +=1
					elif att[0].get('shift_status') == "AB":
						total_absent +=1
					elif att[0].get('shift_status') == "PH":
						total_holiday +=1

				else:
					row1.append('-')
				if att[0].get('shift'):
					row2.append(att[0].get('shift'))
				else:
					row2.append('-')
				in_time = (att[0].get('in_time'))
				out_timee= ''
				formatted_time = ''
				if in_time is not None:
					formatted_time = in_time.strftime('%H:%M')
				else:
					formatted_time = '-'
				out_time = att[0].get('out_time')
				if out_time is not None:
					out_timee = out_time.strftime('%H:%M')
				else:
					out_timee = '-'
				row3.append(formatted_time)
				row4.append(out_timee)
				row5.append(att[0].get('late_entry_hours','-'))
				row6.append(att[0].get('early_exit_hours','-'))
				row7.append(att[0].get('working_hours','-'))
				row8.append(att[0].get('overtime_hours','-'))
			else:
				hh = check_holiday(date,emp.name)
				row1.append(hh)
				row2.append('-')
				row3.append('-')
				row4.append('-')
				row5.append('-')
				row6.append('-')
				row7.append('-')
				row8.append('-')
				if hh == "WH":
					total_weekoff +=1
				elif hh == "BH":
					total_holiday +=1
				elif hh == "DH":
					total_holiday +=1
				elif hh == "PH":
					total_holiday +=1
				elif hh == "FH":
					total_holiday +=1
		row1.append(total_present)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(total_half_day)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append("-")
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append('-')
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(total_absent)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(total_weekoff)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(total_holiday)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(total_lop)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(total_paid_leave)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		
		row1.append('-')
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append('-')
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append('-')
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append('-')
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append('-')
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		row1.append(n_shift)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')
		tw_days=total_present + total_half_day + total_paid_leave +total_holiday
		frappe.errprint(total_present)
		frappe.errprint(total_half_day)
		frappe.errprint(total_paid_leave)
		frappe.errprint(total_holiday)
		row1.append(tw_days)
		row2.append('-')
		row3.append('-')
		row4.append('-')
		row5.append('-')
		row6.append('-')
		row7.append('-')
		row8.append('-')

		data.append(row1)
		data.append(row2)
		data.append(row3)
		data.append(row4)
		data.append(row5)
		data.append(row6)
		data.append(row7)
		data.append(row8)
	return data

def check_holiday(date,emp):
	holiday_list = frappe.db.get_value('Employee',{'name':emp},'holiday_list')
	holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off, `tabHoliday`.others from `tabHoliday List` 
	left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
	doj= frappe.db.get_value("Employee",{'name':emp},"date_of_joining")
	status = ''
	if holiday :
		if doj < holiday[0].holiday_date:
			if holiday[0].weekly_off == 1:
				status = "WH"     
			else:
				if holiday[0].others == "DH":
					status = "DH"
				elif holiday[0].others == "BH":
					status = "BH"
				elif holiday[0].others == "PH":
					status = "PH"
				elif holiday[0].others == "FH":
					status = "FH"
		else:
			status = '*'
	return status
