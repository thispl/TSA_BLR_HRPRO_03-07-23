from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download():
	filename = 'Summary of PF Contribution'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['B'].width = 18 
	ws.column_dimensions['C'].width = 16 
	ws.column_dimensions['D'].width = 16
	ws.column_dimensions['E'].width = 24
	ws.column_dimensions['F'].width = 15

	ws.append(title1(args))


	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10 )
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
	date = datetime.strptime(args['start_date'],'%Y-%m-%d')
	date = date.strftime('%Y')
	data = ["Summary of Labour Welfare Fund deduction by no of employees for the year  " +date]
	
	return data