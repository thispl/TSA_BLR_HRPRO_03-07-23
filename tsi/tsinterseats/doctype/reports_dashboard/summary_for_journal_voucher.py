import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Summary for Journal Voucher Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=29)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=29):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    title_2=title2(args)
    ws.append(title_2)
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=29)
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=29):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=29):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    bold_font = Font(bold=True)

    data_1 = data1(args)
    for data in data_1:
        ws.append(data)
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+3, min_col=1, max_col=29):
        for cell in rows:
            cell.border = border
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
    data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
    return data

@frappe.whitelist()
def title2(args):
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    mon = start_date.strftime('%B %Y').upper()
    data = ['SUMMARY FOR JOURNAL VOUCHER FOR THE MONTH OF %s' % mon]
    return data

@frappe.whitelist()
def header1(args):
    data = ['SI NO.', 'DEPARTMENT', 'HEAD COUNT', 'GL CODE', 'GL DESCRIPTION', 'GROSS SALARY', 'ATTENDANCE BONUS',
            'OVER TIME', 'ARREAR OT', 'DH BONUS', 'SUM OF OVERTIME & ARREAR OT AND DH BONUS', 'SALARY DEDUCTION',
            'NET AMOUNT']
    return data

@frappe.whitelist()
def data1(args):
    data = []
    department_totals = {}

    # Get distinct payroll_cost_center for each department
    department_cost_centers = frappe.get_all('Employee', filters={'status': 'Active'}, group_by='department',
                                             fields=['department', 'payroll_cost_center'])

    for department_cost_center in department_cost_centers:
        department = department_cost_center['department']
        cost_center = department_cost_center['payroll_cost_center']

        # Initialize department_totals if not present
        if department not in department_totals:
            department_totals[department] = {'count': 0, 'gross_pay': 0, 'attendance_bonus': 0,
                                             'declared_holiday_bonus': 0, 'total_deduction': 0, 'net_pay': 0,
                                             'over_time': 0, 'arrear_overtime': 0, 'dh_bonus': 0,
                                             'payroll_cost_center': cost_center}  # Save payroll_cost_center

        # Get the count of employees in the department
        employee_count = frappe.get_value('Employee', {'department': department, 'status': 'Active'}, 'count(name)')

        # Add employee count to department total
        department_totals[department]['count'] += employee_count

        # Get the sum of gross pay, attendance bonus, declared_holiday_bonus, total_deduction, net_pay
        result = frappe.db.sql("""
            SELECT 
                SUM(gross_pay) as gp, 
                SUM(attendance_bonus) as att_bonus, 
                SUM(declared_holiday_bonus) as dh_bonus,
                SUM(total_deduction) as total_deduction,
                SUM(net_pay) as net_pay
            FROM `tabSalary Slip`
            WHERE employee IN (
                SELECT name FROM `tabEmployee` WHERE department = %s AND status = 'Active'
            ) AND start_date >= %s AND end_date <= %s
        """, (department, args.start_date, args.end_date), as_dict=True)

        # Extract the values from the result
        gross_pay = result[0].get('gp') if result and result[0].get('gp') else 0
        attendance_bonus = result[0].get('att_bonus') if result and result[0].get('att_bonus') else 0
        declared_holiday_bonus = result[0].get('dh_bonus') if result and result[0].get('dh_bonus') else 0
        total_deduction = result[0].get('total_deduction') if result and result[0].get('total_deduction') else 0
        net_pay = result[0].get('net_pay') if result and result[0].get('net_pay') else 0

        # Get the sum of over_time and arrear_overtime from the child table
        overtime_result = frappe.db.sql("""
            SELECT 
                SUM(sd.amount) as over_time
            FROM `tabSalary Slip` s
            INNER JOIN `tabSalary Detail` sd ON s.name = sd.parent
            WHERE s.employee IN (
                SELECT name FROM `tabEmployee` WHERE department = %s AND status = 'Active'
            ) AND s.start_date >= %s AND s.end_date <= %s
                AND sd.salary_component = 'Over Time'
        """, (department, args.start_date, args.end_date), as_dict=True)

        arrear_overtime_result = frappe.db.sql("""
            SELECT 
                SUM(sd.amount) as arrear_overtime
            FROM `tabSalary Slip` s
            INNER JOIN `tabSalary Detail` sd ON s.name = sd.parent
            WHERE s.employee IN (
                SELECT name FROM `tabEmployee` WHERE department = %s AND status = 'Active'
            ) AND s.start_date >= %s AND s.end_date <= %s
                AND sd.salary_component = 'Arrear Overtime'
        """, (department, args.start_date, args.end_date), as_dict=True)

        # Extract the values from the child table results
        over_time = overtime_result[0].get('over_time') if overtime_result and overtime_result[0].get('over_time') else 0
        arrear_overtime = arrear_overtime_result[0].get('arrear_overtime') if arrear_overtime_result and arrear_overtime_result[0].get('arrear_overtime') else 0
        declared_holiday_bonus = result[0].get('dh_bonus') if result and result[0].get('dh_bonus') else 0

        # Add values to department total
        department_totals[department]['gross_pay'] += gross_pay
        department_totals[department]['attendance_bonus'] += attendance_bonus
        department_totals[department]['over_time'] += over_time
        department_totals[department]['arrear_overtime'] += arrear_overtime
        department_totals[department]['declared_holiday_bonus'] += declared_holiday_bonus
        department_totals[department]['total_deduction'] += total_deduction
        department_totals[department]['net_pay'] += net_pay

    # Iterate through department_totals and create the final data list
    for index, (department, values) in enumerate(department_totals.items(), start=1):
        sum_overtime_arrear_dh_bonus = values['over_time'] + values['arrear_overtime'] + values['declared_holiday_bonus']
        data.append([
            index,
            department,
            values['count'],
            values['payroll_cost_center'],  
            department,
            values['gross_pay'], 
            values['attendance_bonus'], 
            values['over_time'],
            values['arrear_overtime'],
            values['declared_holiday_bonus'],
            sum_overtime_arrear_dh_bonus,  # Sum of OVER TIME, ARREAR OT, and DH BONUS
            values['total_deduction'],
            values['net_pay'], 
              # Add other columns as needed
        ])

    return data




        
       










