import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
	filename = 'Salary Summary Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	title_1=title1(args)
	ws.append(title_1)
	ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=25)
	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=25):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	title_2=title2(args)
	ws.append(title_2)
	ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=25)
	for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=25):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	header_1=header1(args)
	ws.append(header_1)
	bold_font = Font(bold=True)
	for cell in ws["3:3"]:
		cell.font = bold_font
	for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=25):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')
	
	bold_font = Font(bold=True)
	data_1 = data1(args)
	for data in data_1:
		ws.append(data)
	title_3=title3(args)
	ws.append(title_3)
	title_4=title4(args)
	ws.append(title_4)
	header_2=header2(args)
	ws.append(header_2)
	data_2 = data2(args)
	for fdata in data_2:
		ws.append(fdata)
	title_5=title5(args)
	ws.append(title_5)
	title_6=title6(args)
	ws.append(title_6)
	header_3=header3(args)
	ws.append(header_3)
	data_3 = data3(args)
	for adata in data_3:
		ws.append(adata)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
	data = ['TS INTERSEATS INDIA PVT LTD']
	return data
@frappe.whitelist()
def title2(args):
	month = datetime.strptime(str(args.start_date),'%Y-%m-%d')
	mon = str(month.strftime('%b') +' '+ str(month.strftime('%Y')))
	data = ['SUMMARY OF WORKERS SALARY  FOR THE MONTH OF %s'%mon]
	return data
@frappe.whitelist()
def header1(args):
	data = ['Sr.No.','Departments','Basic Salary','Dearness Allowance','Arr.Basic Salary & DA','Special Allowance','Incentives','Arrears Salary','Over Time Salary','Arr. Over Time Salary','Shift Allowance','Gross Salary','PF','PT','ESI','LWF','Salary Deduction','Deductions','Attd. Bonus','Casual Leave Encashment','Earned Leave Encashment','Extra Pay','Bonus Yearly','Net Salary','Remarks']
	return data
@frappe.whitelist()
def data1(args):
	data = []

	sa = frappe.get_all('Salary Slip', {'salary_structure': "Worker"}, ['department'])

	department_totals = {}
	
	for i in sa:
		department = i.department
		department_escaped = frappe.db.escape(department)

		def get_salary_component_total(salary_component):
			return frappe.db.sql(f"""
				select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
				left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
				where `tabSalary Slip`.department = {department_escaped} and
					`tabSalary Detail`.salary_component = '{salary_component}' and
					`tabSalary Slip`.start_date >= '{args.start_date}' and
					`tabSalary Slip`.end_date <= '{args.end_date}'
			""", as_dict=True)[0].total or 0

		b = get_salary_component_total('Basic')
		da = get_salary_component_total('Dearness Allowance')
		arb = get_salary_component_total('Arrear Basic')
		sa = get_salary_component_total('Special Allowance')
		i = get_salary_component_total('Incentives')
		a = get_salary_component_total('Arrear')
		ot = get_salary_component_total('Overtime')
		aot = get_salary_component_total('Arrear Overtime')
		sha = get_salary_component_total('Shift Allowance')

		gross = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].gross

		pf = get_salary_component_total('Provident Fund')
		pt = get_salary_component_total('Professional Tax')
		esi = get_salary_component_total('Employee State Insurance')
		lwf = get_salary_component_total('LWF')
		sd = get_salary_component_total('salary Deduction')

		deduction = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.total_deduction) as deduction from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].deduction

		ab = get_salary_component_total('Attendance Bonus')
		cle = get_salary_component_total('Causal Leave Encashment')
		ele = get_salary_component_total('Earned Leave Encashment')
		ep = get_salary_component_total('Extra Pay')
		pb = get_salary_component_total('Position Bonus')

		npay = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.net_pay) as npay from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].npay

		department_totals[department] = [b, da, arb, sa, i, a, ot, aot, sha, gross, pf, pt, esi, lwf, sd, deduction, ab, cle, ele, ep, pb, npay]

	total_basic = sum(x[0] for x in department_totals.values())
	total_da = sum(x[1] for x in department_totals.values())
	total_arb = sum(x[2] for x in department_totals.values())
	total_sa = sum(x[3] for x in department_totals.values())
	total_i = sum(x[4] for x in department_totals.values())
	total_a = sum(x[5] for x in department_totals.values())
	total_ot = sum(x[6] for x in department_totals.values())
	total_aot = sum(x[7] for x in department_totals.values())
	total_sha = sum(x[8] for x in department_totals.values())
	total_gross = sum(x[9] for x in department_totals.values())
	total_pf = sum(x[10] for x in department_totals.values())
	total_pt = sum(x[11] for x in department_totals.values())
	total_esi = sum(x[12] for x in department_totals.values())
	total_lwf = sum(x[13] for x in department_totals.values())
	total_sd = sum(x[14] for x in department_totals.values())
	total_deduction = sum(x[15] for x in department_totals.values())
	total_ab = sum(x[16] for x in department_totals.values())
	total_cle = sum(x[17] for x in department_totals.values())
	total_ele = sum(x[18] for x in department_totals.values())
	total_ep = sum(x[19] for x in department_totals.values())
	total_pb = sum(x[20] for x in department_totals.values())
	total_npay = sum(x[21] for x in department_totals.values())

	for index, (department, totals) in enumerate(department_totals.items(), start=1):
		data.append([index, department] + totals)

	row = ['', 'Grand Total', total_basic, total_da, total_arb, total_sa, total_i, total_a, total_ot, total_aot, total_sha, total_gross, total_pf, total_pt, total_esi, total_lwf, total_sd, total_deduction, total_ab, total_cle, total_ele, total_ep, total_pb, total_npay]
	data.append(row)

	return data
@frappe.whitelist()
def title3(args):
	data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
	return data
@frappe.whitelist()
def title4(args):
	month = datetime.strptime(str(args.start_date),'%Y-%m-%d')
	mon = str(month.strftime('%b') +' '+ str(month.strftime('%Y')))
	data = ['SUMMARY OF FACTORY STAFF SALARY  FOR THE MONTH OF %s'%mon]
	return data
@frappe.whitelist()
def header2(args):
	data = ['Sr.No.','Departments','Basic Salary','Dearness Allowance','Arr.Basic Salary & DA','House Rent Allowance','Transport Allowance','Special Allowance','Medical Allowance','Position Bonus','Incentives','Over Time Salary','Arr. Over Time Salary','Arrears Salary','Shift Allowance','Gross Salary','PF','PT','ESI','INCENTE','TDS Deduction','LWF','25% Deduction','Deductions','Attendance Bonus','Casual Leave Encashment','Earned Leave Encashment','Bonus Yearly','Net Salary','Extra Holiday Bonus']
	return data
@frappe.whitelist()
def data2(args):
	data = []

	sa = frappe.get_all('Salary Slip', {'salary_structure': "Factory Staff"}, ['department'])

	department_totals = {}
	
	for i in sa:
		department = i.department
		department_escaped = frappe.db.escape(department)

		def get_salary_component_total(salary_component):
			return frappe.db.sql(f"""
				select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
				left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
				where `tabSalary Slip`.department = {department_escaped} and
					`tabSalary Detail`.salary_component = '{salary_component}' and
					`tabSalary Slip`.start_date >= '{args.start_date}' and
					`tabSalary Slip`.end_date <= '{args.end_date}'
			""", as_dict=True)[0].total or 0

		b = get_salary_component_total('Basic')
		da = get_salary_component_total('Dearness Allowance')
		
		arb = get_salary_component_total('Arrear Basic')
		hra = get_salary_component_total('House Rent Allowance')
		ta = get_salary_component_total('Transport Allowance')
		sa = get_salary_component_total('Special Allowance')
		ma = get_salary_component_total('Medical Allowance')
		pb = get_salary_component_total('Position Bonus')

		i = get_salary_component_total('Incentives')
		ot = get_salary_component_total('Overtime')
	   
		aot = get_salary_component_total('Arrear Overtime')
		a = get_salary_component_total('Arrear')

		sha = get_salary_component_total('Shift Allowance')

		gross = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].gross

		pf = get_salary_component_total('Provident Fund')
		pt = get_salary_component_total('Professional Tax')
		esi = get_salary_component_total('Employee State Insurance')
		ic = get_salary_component_total('Incente')
		tds = get_salary_component_total('TDS')

		lwf = get_salary_component_total('LWF')
		d = get_salary_component_total('25% Deduction')

		deduction = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.total_deduction) as deduction from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].deduction

		ab = get_salary_component_total('Attendance Bonus')
		cle = get_salary_component_total('Causal Leave Encashment')
		ele = get_salary_component_total('Earned Leave Encashment')
		b = get_salary_component_total(' Bonus')

		npay = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.net_pay) as npay from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].npay
		hb = get_salary_component_total('Holiday Bonus')
		department_totals[department] = [b, da, arb,hra,ta, sa,ma,pb, i,  ot, aot,a, sha, gross, pf, pt, esi,ic,tds, lwf,d, deduction, ab, cle, ele,b,npay,hb]

	total_basic = sum(x[0] for x in department_totals.values())
	total_da = sum(x[1] for x in department_totals.values())
	total_arb = sum(x[2] for x in department_totals.values())
	total_hra = sum(x[3] for x in department_totals.values())
	total_ta = sum(x[4] for x in department_totals.values())
	total_sa = sum(x[5] for x in department_totals.values())
	total_ma = sum(x[6] for x in department_totals.values())
	total_pb = sum(x[7] for x in department_totals.values())
	total_i = sum(x[8] for x in department_totals.values())
	total_ot = sum(x[9] for x in department_totals.values())
	total_aot = sum(x[10] for x in department_totals.values())
	total_a = sum(x[11] for x in department_totals.values())
	total_sha = sum(x[12] for x in department_totals.values())
	total_gross = sum(x[13] for x in department_totals.values())
	total_pf = sum(x[14] for x in department_totals.values())
	total_pt = sum(x[15] for x in department_totals.values())
	total_esi = sum(x[16] for x in department_totals.values())
	total_lwf = sum(x[17] for x in department_totals.values())
	total_d = sum(x[18] for x in department_totals.values())
	total_deduction = sum(x[19] for x in department_totals.values())
	total_ab = sum(x[20] for x in department_totals.values())
	total_cle = sum(x[21] for x in department_totals.values())
	total_ele = sum(x[22] for x in department_totals.values())
	total_b = sum(x[23] for x in department_totals.values())
	total_npay = sum(x[24] for x in department_totals.values())
	total_hb = sum(x[25] for x in department_totals.values())

	for index, (department, totals) in enumerate(department_totals.items(), start=1):
		data.append([index, department] + totals)

	row = ['', 'Grand Total', total_basic, total_da, total_arb, total_hra, total_ta, total_sa, total_ma, total_pb, total_i, total_ot, total_aot, total_a, total_sha, total_gross, total_pf, total_pt, total_esi, total_lwf, total_d, total_deduction, total_ab, total_cle, total_ele, total_b, total_npay, total_hb]
	data.append(row)

	return data

@frappe.whitelist()
def title5(args):
	data = ['TS INTERSEATS INDIA PVT LTD']
	return data
@frappe.whitelist()
def title6(args):
	month = datetime.strptime(str(args.start_date),'%Y-%m-%d')
	mon = str(month.strftime('%b') +' '+ str(month.strftime('%Y')))
	data = ['SUMMARY OF ADMINISTRATION STAFF SALARY  FOR THE MONTH OF %s'%mon]
	return data
@frappe.whitelist()
def header3(args):
	data = ['Sr.No.','Departments','Basic Salary','Dearness Allowance','Arr.Basic Salary & DA','House Rent Allowance','Transport Allowance','Special Allowance','Medical Allowance','Position Bonus','Incentives','Over Time Salary','Arr. Over Time Salary','Arrears Salary','Shift Allowance','Gross Salary','PF','PT','ESI','INCENTE','TDS Deduction','LWF','25% Deduction','Deductions','Attendance Bonus','Casual Leave Encashment','Earned Leave Encashment','Bonus Yearly','Net Salary','Extra Holiday Bonus']
	return data
@frappe.whitelist()
def data3(args):
	data = []

	sa = frappe.get_all('Salary Slip', {'salary_structure': "Administration Staff"}, ['department'])

	department_totals = {}
	
	for i in sa:
		department = i.department
		department_escaped = frappe.db.escape(department)

		def get_salary_component_total(salary_component):
			return frappe.db.sql(f"""
				select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
				left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
				where `tabSalary Slip`.department = {department_escaped} and
					`tabSalary Detail`.salary_component = '{salary_component}' and
					`tabSalary Slip`.start_date >= '{args.start_date}' and
					`tabSalary Slip`.end_date <= '{args.end_date}'
			""", as_dict=True)[0].total or 0

		b = get_salary_component_total('Basic')
		da = get_salary_component_total('Dearness Allowance')
		
		arb = get_salary_component_total('Arrear Basic')
		hra = get_salary_component_total('House Rent Allowance')
		ta = get_salary_component_total('Transport Allowance')
		sa = get_salary_component_total('Special Allowance')
		ma = get_salary_component_total('Medical Allowance')
		pb = get_salary_component_total('Position Bonus')

		i = get_salary_component_total('Incentives')
		ot = get_salary_component_total('Overtime')
	   
		aot = get_salary_component_total('Arrear Overtime')
		a = get_salary_component_total('Arrear')

		sha = get_salary_component_total('Shift Allowance')

		gross = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].gross

		pf = get_salary_component_total('Provident Fund')
		pt = get_salary_component_total('Professional Tax')
		esi = get_salary_component_total('Employee State Insurance')
		ic = get_salary_component_total('Incente')
		tds = get_salary_component_total('TDS')

		lwf = get_salary_component_total('LWF')
		d = get_salary_component_total('25% Deduction')

		deduction = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.total_deduction) as deduction from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].deduction

		ab = get_salary_component_total('Attendance Bonus')
		cle = get_salary_component_total('Causal Leave Encashment')
		ele = get_salary_component_total('Earned Leave Encashment')
		b = get_salary_component_total(' Bonus')

		npay = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.net_pay) as npay from `tabSalary Slip`
			where `tabSalary Slip`.department = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].npay
		hb = get_salary_component_total('Holiday Bonus')
		department_totals[department] = [b, da, arb,hra,ta, sa,ma,pb, i,  ot, aot,a, sha, gross, pf, pt, esi,ic,tds, lwf,d, deduction, ab, cle, ele,b,npay,hb]

	total_basic = sum(x[0] for x in department_totals.values())
	total_da = sum(x[1] for x in department_totals.values())
	total_arb = sum(x[2] for x in department_totals.values())
	total_hra = sum(x[3] for x in department_totals.values())
	total_ta = sum(x[4] for x in department_totals.values())
	total_sa = sum(x[5] for x in department_totals.values())
	total_ma = sum(x[6] for x in department_totals.values())
	total_pb = sum(x[7] for x in department_totals.values())
	total_i = sum(x[8] for x in department_totals.values())
	total_ot = sum(x[9] for x in department_totals.values())
	total_aot = sum(x[10] for x in department_totals.values())
	total_a = sum(x[11] for x in department_totals.values())
	total_sha = sum(x[12] for x in department_totals.values())
	total_gross = sum(x[13] for x in department_totals.values())
	total_pf = sum(x[14] for x in department_totals.values())
	total_pt = sum(x[15] for x in department_totals.values())
	total_esi = sum(x[16] for x in department_totals.values())
	total_lwf = sum(x[17] for x in department_totals.values())
	total_d = sum(x[18] for x in department_totals.values())
	total_deduction = sum(x[19] for x in department_totals.values())
	total_ab = sum(x[20] for x in department_totals.values())
	total_cle = sum(x[21] for x in department_totals.values())
	total_ele = sum(x[22] for x in department_totals.values())
	total_b = sum(x[23] for x in department_totals.values())
	total_npay = sum(x[24] for x in department_totals.values())
	total_hb = sum(x[25] for x in department_totals.values())

	for index, (department, totals) in enumerate(department_totals.items(), start=1):
		data.append([index, department] + totals)

	row = ['', 'Grand Total', total_basic, total_da, total_arb, total_hra, total_ta, total_sa, total_ma, total_pb, total_i, total_ot, total_aot, total_a, total_sha, total_gross, total_pf, total_pt, total_esi, total_lwf, total_d, total_deduction, total_ab, total_cle, total_ele, total_b, total_npay, total_hb]
	data.append(row)

	return data
@frappe.whitelist()
def title7(args):
	data = ['TS INTERSEATS INDIA PVT LTD']
	return data
@frappe.whitelist()
def title8(args):
	month = datetime.strptime(str(args.start_date),'%Y-%m-%d')
	mon = str(month.strftime('%b') +' '+ str(month.strftime('%Y')))
	data = ['SUMMARY OF WORKERS SALARY  FOR THE MONTH OF %s'%mon]
	return data
@frappe.whitelist()
def header4(args):
	data = ['Sr.No.','Departments','No Of Employee','Basic Salary','Dearness Allowance','Arr.Basic Salary & DA','Special Allowance','Incentives','Arrears Salary','Over Time Salary','Arr. Over Time Salary','Shift Allowance','Gross Salary','PF','PT','ESI','LWF','Salary Deduction','Deductions','Attd. Bonus','Casual Leave Encashment','Earned Leave Encashment','Extra Pay','Bonus Yearly','Net Salary','Remarks']
	return data
@frappe.whitelist()
def data4(args):
	data = []

	sa = frappe.get_all('Salary Slip', {'salary_structure': "Administration Staff"}, ['department'])