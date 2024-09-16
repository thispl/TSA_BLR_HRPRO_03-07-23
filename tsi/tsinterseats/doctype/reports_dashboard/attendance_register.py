import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download(start_date,end_date,department=None):
    filename = 'Attendance Register'
    args = {'start_date':start_date,'end_date':end_date,'department':department}
    frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
    enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)
   
def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 13 
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 20 
    header_date = title (args)
    ws.append(header_date) 

    header_date = title1(args)
    ws.append(title1(args))
    
    row3=(get_dep(args))
    ws.append(row3)

    row4=(get_date(args))
    ws.append(row4)

    row5=(get_day(args))
    ws.append(row5)

    emp= get_data(args)
    for e in emp:
        ws.append(e)

    ws.append([''])

    abbr = ['']
    ws.append(abbr)

    tsi=['','TS Interseats India Private Limited'," "," "," "," ","10","DH/PN",'Night Shift Present on Declared Holiday',"","","PREPARED BY","","","","","","CHECKED BY","","","","","","APPROVED BY"]
    tsi1=['','Abbreviation for Attendance'," "," "," "," ","11","DH/PP",'Partial Present on Declared Holiday',"","","","","","",""]
    tsi2=['','SI No',"Mark","Description"," "," ","12","P/CL",'Half Day Casual Leave',"","","","","","",""]
    tsi3=['','1',"P","Present"," "," ","13","P/EL",'Half Day Earned Holiday',"","","","","","",""]
    tsi4=['','2',"P/N","Present Night Shift"," "," ","14","AP",'Assumed Present',"","","","","","",""]
    tsi5=['','3',"AB","Absent"," "," ","15","ESI",'ESI Leave',"","","","","","",""]
    tsi6=['','4',"DH","Declared Holiday"," "," ","16","CO",'Complimentary Off',"","","","","","",""]
    tsi7=['','5',"WH","Weekly Holiday"," "," ","17","CL",'Casual Leave',"","","","","","",""]
    tsi8=['','6',"PH","Public Holiday"," "," ","18","EL",'Earned Leave',"","","Mr.Chandrappa K M","","","","","","Mr.Chandrahasa ","","","","","","Mrs.Pornprapha Pholdahan"]
    tsi9=['','7',"FH","Festival Leave"," "," ","19","LOP",'Loss of Pay',"","","","","","",""]
    tsi10=['','8',"BH","Block Holiday"," "," ","20","P/AB",'Half Day Absent',"","","(EXECUTIVE - HR)","","","","","","(MANAGER - HR)","","","","","","(CFO)"]
    tsi11=['','9',"DH/P","Present on Declared Holiday"," "," ","21","*",'No Attendance (for New Joining)',"","","","","","",""]


    
    ws.append(tsi)
    ws.append(tsi1)
    ws.append(tsi2)
    ws.append(tsi3)
    ws.append(tsi4)
    ws.append(tsi5)
    ws.append(tsi6)
    ws.append(tsi7)
    ws.append(tsi8)
    ws.append(tsi9)
    ws.append(tsi10)
    ws.append(tsi11)

    
    

    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(get_dep(args)) )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(get_dep(args)) )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)), end_row=4, end_column=len(get_dep(args)) )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 1, end_row=4, end_column=len(get_dep(args)) - 1 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 2, end_row=4, end_column=len(get_dep(args)) - 2 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 3, end_row=4, end_column=len(get_dep(args)) - 3 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 4, end_row=4, end_column=len(get_dep(args)) - 4 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 5, end_row=4, end_column=len(get_dep(args)) - 5 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 6, end_row=4, end_column=len(get_dep(args)) - 6 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 7, end_row=4, end_column=len(get_dep(args)) - 7 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 8, end_row=4, end_column=len(get_dep(args)) - 8 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 9, end_row=4, end_column=len(get_dep(args)) - 9 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 10, end_row=4, end_column=len(get_dep(args)) - 10 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 11, end_row=4, end_column=len(get_dep(args)) - 11 )
    ws.merge_cells(start_row=3, start_column=len(get_dep(args)) - 12, end_row=4, end_column=len(get_dep(args)) - 12 )
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=len(get_dep(args)) - 13 )
    ws.merge_cells(start_row=len(get_data(args))+5, start_column=1, end_row=len(get_data(args))+5, end_column=len(get_dep(args)) - 13 )
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2 )
    ws.merge_cells(start_row=3, start_column=3, end_row=4, end_column=3 )
    ws.merge_cells(start_row=3, start_column=4, end_row=4, end_column=4 )
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1 )
    ws.merge_cells(start_row=len(get_data(args))+8, start_column=2, end_row=len(get_data(args))+8, end_column=4 )
    ws.merge_cells(start_row=len(get_data(args))+9, start_column=2, end_row=len(get_data(args))+9, end_column=4 )
    ws.merge_cells(start_row=len(get_data(args))+8, start_column=12, end_row=len(get_data(args))+9, end_column=15 )
    ws.merge_cells(start_row=len(get_data(args))+10, start_column=12, end_row=len(get_data(args))+15, end_column=15 )
    ws.merge_cells(start_row=len(get_data(args))+16, start_column=12, end_row=len(get_data(args))+17, end_column=15 )
    ws.merge_cells(start_row=len(get_data(args))+18, start_column=12, end_row=len(get_data(args))+19, end_column=15 )
    ws.merge_cells(start_row=len(get_data(args))+8, start_column=18, end_row=len(get_data(args))+9, end_column=21 )
    ws.merge_cells(start_row=len(get_data(args))+10, start_column=18, end_row=len(get_data(args))+15, end_column=21 )
    ws.merge_cells(start_row=len(get_data(args))+16, start_column=18, end_row=len(get_data(args))+17, end_column=21 )
    ws.merge_cells(start_row=len(get_data(args))+18, start_column=18, end_row=len(get_data(args))+19, end_column=21 )
    ws.merge_cells(start_row=len(get_data(args))+8, start_column=24, end_row=len(get_data(args))+9, end_column=27 )
    ws.merge_cells(start_row=len(get_data(args))+10, start_column=24, end_row=len(get_data(args))+15, end_column=27 )
    ws.merge_cells(start_row=len(get_data(args))+16, start_column=24, end_row=len(get_data(args))+17, end_column=27 )
    ws.merge_cells(start_row=len(get_data(args))+18, start_column=24, end_row=len(get_data(args))+19, end_column=27 )	
    
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center',vertical='center')
    for header in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=len(get_dep(args))):
            for cell in header:
                cell.fill = PatternFill(fgColor='002366', fill_type = "solid")
                cell.font = Font(bold=True,color='FFFFFF',size=10)
                cell.alignment = align_center

    for header in ws.iter_rows(min_row=3, min_col=5, max_row=3, max_col=len(get_dep(args)) - 13 ):
            for cell in header:
                cell.fill = PatternFill(fgColor='002366', fill_type = "solid")
                cell.font = Font(bold=True,color='002366',size=10)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=6, min_col=5, max_row=len(get_data(args))+6, max_col=len(get_dep(args))):
            for cell in header:
                cell.alignment = align_center
                
    for header in ws.iter_rows(min_row=6, min_col=5, max_row=len(get_data(args))+5, max_col=len(get_dep(args))-13):
            for cell in header:
                cell.alignment = align_center
                if cell.value == "WH":
                    cell.fill = PatternFill(fgColor='DAF7A6', fill_type = "solid")
                if cell.value == "CL":
                    cell.fill = PatternFill(fgColor='5BDE00', fill_type = "solid")
                if cell.value == "ESI":
                    cell.fill = PatternFill(fgColor='F6DDCC', fill_type = "solid")
                if cell.value == "P/CL":
                    cell.fill = PatternFill(fgColor='84C370', fill_type = "solid")
                if cell.value == "EL":
                    cell.fill = PatternFill(fgColor='79CFD6', fill_type = "solid")
    for header in ws.iter_rows(min_row=len(get_data(args))+5, min_col=len(get_dep(args))-13, max_row=len(get_data(args))+6, max_col=len(get_dep(args))):
            for cell in header:
                cell.alignment = align_center
                cell.font = Font(bold=True)

    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=2, max_row=len(get_data(args))+10, max_col=4):
        for cell in header:
            cell.font = Font(bold=True)
    
    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=2, max_row=len(get_data(args))+19, max_col=4):
        for cell in header:
            cell.border = border_thin		

    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=7, max_row=len(get_data(args))+19, max_col=9):
        for cell in header:
            cell.border = border_thin
        
    for header in ws.iter_rows(min_row=len(get_data(args))+10, min_col=2, max_row=len(get_data(args))+19, max_col=3):
        for cell in header:
            cell.alignment = align_center

    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=7, max_row=len(get_data(args))+19, max_col=8):
        for cell in header:
            cell.alignment = align_center

    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=12, max_row=len(get_data(args))+9, max_col=15):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)
    for header in ws.iter_rows(min_row=len(get_data(args))+16, min_col=12, max_row=len(get_data(args))+17, max_col=15):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)

    for header in ws.iter_rows(min_row=len(get_data(args))+18, min_col=12, max_row=len(get_data(args))+19, max_col=15):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)

    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=18, max_row=len(get_data(args))+9, max_col=21):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)
    for header in ws.iter_rows(min_row=len(get_data(args))+16, min_col=18, max_row=len(get_data(args))+17, max_col=21):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)

    for header in ws.iter_rows(min_row=len(get_data(args))+18, min_col=24, max_row=len(get_data(args))+19, max_col=27):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)

    for header in ws.iter_rows(min_row=len(get_data(args))+8, min_col=24, max_row=len(get_data(args))+9, max_col=27):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)
    for header in ws.iter_rows(min_row=len(get_data(args))+16, min_col=24, max_row=len(get_data(args))+17, max_col=27):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)

    for header in ws.iter_rows(min_row=len(get_data(args))+18, min_col=18, max_row=len(get_data(args))+19, max_col=21):
        for cell in header:
            cell.alignment = align_center
            cell.font = Font(bold=True,size=14)

    header_range = ws['A1':ws.cell(row=len(get_data(args))+5, column=len(get_dep(args))).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin

    # for header in ws.iter_rows(min_row=6, min_col=5, max_row=len(get_data(args)), max_col=len(get_dep(args))-13):
    # 		for cell in header:
    # 			cell.alignment = align_center
    
    ws.freeze_panes = 'E5'

    ws.sheet_view.zoomScale = 100
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    frappe.log_error(title='test',message=xlsx_file)
    return xlsx_file    

def build_xlsx_response(filename,args):
    frappe.log_error(title='test',message=filename)
    xlsx_file = make_xlsx(filename,args)
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": ' ',
            "attached_to_doctype": 'Reports Dashboard',
            "attached_to_field": 'attach',
            "file_name": filename + '.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    frappe.db.commit()
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('Reports Dashboard',None,'attach',attached_file.file_url)


# def build_xlsx_response(filename):
# 	xlsx_file = make_xlsx(filename)
# 	frappe.response['filename'] = filename + '.xlsx'
# 	frappe.response['filecontent'] = xlsx_file.getvalue()
# 	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title(args):
    data = ["TS Interseats India Private Limited"]
    return data

@frappe.whitelist()
def title1(args):
    date = datetime.strptime(args['start_date'],'%Y-%m-%d')
    date = date.strftime('%B-%Y')
    data = ["Worker Attendance Sheet " +date]
    
    return data


@frappe.whitelist()
def get_data(args):
    data = []
    row = []
    

    if args['department']:
        emp = frappe.get_all("Employee",{'department':args['department'],'status':'Active'},['*'],order_by='name ASC')	
    else:
        emp = frappe.get_all("Employee",{'status':'Active'},['*'],order_by='name ASC')
    i=1
    tdcnt =0 
    tpds = 0
    twh=0
    tbh=0
    tdh=0
    tph=0
    tfh=0
    tcl=0
    tel=0
    tesi=0
    tlop=0
    ttpd=0
    for e in emp:
        row=[i,]
        row += [e.name,e.employee_name,e.department]
        dates=get_dates(args)
        cnt = 0
        pd=0
        wh=0
        bh=0
        dh=0
        ph=0
        fh=0
        cl=0
        el=0
        esi=0
        lop=0
        tpd=0

        for date in dates:
            if frappe.db.exists("Attendance",{'attendance_date':date,'employee':e.name,'docstatus':('!=','2')}):
                att = frappe.db.get_value("Attendance",{'attendance_date':date,'employee':e.name},['shift_status'])
                if att:
                    row.append(att)
                    if att =="P":
                        pd += 1
                        # cell.font = Font(bold=True,size=14)
                    elif att == "P/N":
                        pd +=1
                    elif att == "P/AB":
                        pd +=0.5
                    elif att == "DH/P":
                        pd +=1
                    elif att == "DH/PN":
                        pd +=1
                    elif att == "DH/PP":
                        pd +=0.5
                    elif att == "CL":
                        cl +=1
                    elif att == "P/CL":
                        cl +=0.5
                        pd +=0.5
                    elif att == "EL":
                        el +=1
                    elif att == "P/EL":
                        el +=0.5
                        pd +=0.5
                    elif att == "ESI":
                        esi +=1
                    elif att == "LOP":
                        lop +=1

                else:
                    row.append('-')
            else:
                hh = check_holiday(date,e.name)
                row.append(hh)
                if hh == "WH":
                    wh+=1
                elif hh == "BH":
                    bh +=1
                elif hh == "DH":
                    dh +=1
                elif hh == "PH":
                    ph +=1
                elif hh == "FH":
                    fh +=1
            cnt += 1
        row.append(cnt)
        row.append(pd)
        row.append(bh)
        row.append(dh)
        row.append(ph)
        row.append(fh)
        row.append(wh)
        row.append(cl)
        row.append(el)
        row.append(esi)
        row.append(lop)
        tpd=pd+bh+dh+ph+fh+wh+cl+el
        row.append(tpd)
        row.append("")
        data.append(row)
        i+=1
        tdcnt +=cnt
        tpds +=pd
        tbh +=bh
        tdh +=dh
        tph +=ph
        tfh +=fh
        twh +=wh
        tcl +=cl
        tel +=el
        tesi +=esi
        tlop +=lop
        ttpd +=tpd

    row1 = ['','','','']
    dates = get_dates(args)
    for date in dates:
        row1.extend([''])

    row1.append(tdcnt)
    row1.append(tpds)
    row1.append(tbh)
    row1.append(tdh)
    row1.append(tph)
    row1.append(tfh)
    row1.append(twh)
    row1.append(tcl)
    row1.append(tel)
    row1.append(tesi)
    row1.append(tlop)
    row1.append(ttpd)
    data.append(row1)
    return data

# def get_lastrow(args):
# 	row=[""]
# 	last_row=get_dates(args)
# 	count=0
# 	for i in last_row:
# 		count+=1
# 	row.append(count)
# 	return row



@frappe.whitelist()
def get_dep(args):
    data = []
    data += ['S No','Employee No','Name of Employee','Department']
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        date = date.strftime('%d')
        data.extend([date])
    data.extend(["Total Days","Present days","BH","DH","PH","FH","WH","CL","EL","ESI","Loss Of Pay","Total Paid day","Remarks"])
    return data

@frappe.whitelist()
def get_day(args):
    data = []
    data += ['','','','']
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        date = date.strftime('%a').upper()
        data.extend([date])
    return data

@frappe.whitelist()
def get_date(args):
    data = []
    data += ["","","",""]
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        date = date.strftime('%d')
        data.extend([date])
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args['end_date'], 1), args['start_date'])
    dates = [add_days(args['start_date'], i) for i in range(0, no_of_days)]
    return dates

def check_holiday(date,e):
    holiday_list = frappe.db.get_value('Employee',{'name':e},'holiday_list')
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off, `tabHoliday`.others from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
    doj= frappe.db.get_value("Employee",{'name':e},"date_of_joining")
    status = ''
    if holiday :
        if doj < holiday[0].holiday_date:
            if holiday[0].weekly_off == 1:
                status = "WH"     
            else:
                if holiday[0].others == "DH":
                    status = "DH"
                elif holiday[0].others == "BH":
                    status = "BH"
                elif holiday[0].others == "PH":
                    status = "PH"
                elif holiday[0].others == "FH":
                    status = "FH"
        else:
            status = '*'
    return status


