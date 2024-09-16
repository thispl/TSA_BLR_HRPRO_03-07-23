import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download(start_date,end_date,department=None,category=None):
    filename = 'Salary Summary Register'
    args = {'start_date':start_date,'end_date':end_date,'department':department,'category':category}
    frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
    enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)
   
def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25

    ws.append(title(args))
    ws.append(title1(args))
    ws.append([''])
    ws.append(['','','','','','','','','','',"STANDARD SALARY",'','','','','','','','EARNINGS','','','','','','','','','','','','','','DEDUCTIONS'])
    ws.append(['Sl No.','Employee No','DATE OF JOINING','Employee Name','Paid Days','Over Time (Hrs)','Arrear Days','Arr.Over Time (Hrs)','Salary Deduction','Shift Days','Basic+DA Salary','DA','HRA','Conveyance Allowance','Medical Allowance','Special  Allowance','Position Bonus','Gross Salary','Basic+DA Salary','DA','Arr. Basic Salary','HRA','Conveyance Allowance','Medical Allowance','Special  Allowance','Position Bonus','Incentives','Over Time','Arrears Over Time','Arrears Salary','Shift allowance','Gross Salary','PF','PT','ESI','IT Deduction','LWF','Salary Deduction','Total Deduction','Attendance Bonus','Casual Leave Encashment','Earned Leave Encashment','Bonus Yearly','Net Salary','Declared Holiday Bonus','Department','Cost Center','Cost Center Name'])
    emp= get_data(args)
    for e in emp:
        ws.append(e)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=48 )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=48 )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=48 )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=10 )
    ws.merge_cells(start_row=4, start_column=11, end_row=4, end_column=18 )
    ws.merge_cells(start_row=4, start_column=19, end_row=4, end_column=32 )
    ws.merge_cells(start_row=4, start_column=33, end_row=4, end_column=39 )
    ws.merge_cells(start_row=4, start_column=40, end_row=4, end_column=48 )
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    border_thick = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick'))
    align_center = Alignment(horizontal='center',vertical='center')
    for header in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=48):
        for cell in header:
            cell.font = Font(bold=True)
    for header in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=48):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=len(get_data(args))+5, max_row=len(get_data(args))+5, min_col=1, max_col=48):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    header_range = ws['A1':ws.cell(row=len(get_data(args))+5, column=48).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    frappe.log_error(title='test',message=xlsx_file)
    return xlsx_file    

def build_xlsx_response(filename,args):
    frappe.log_error(title='test',message=filename)
    xlsx_file = make_xlsx(filename,args)
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": ' ',
            "attached_to_doctype": 'Reports Dashboard',
            "attached_to_field": 'salary_summary_register',
            "file_name": filename + '.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    frappe.db.commit()
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('Reports Dashboard',None,'salary_summary_register',attached_file.file_url)

@frappe.whitelist()
def title(args):
    data = ["TS INTERSEATS INDIA PVT LTD"]
    return data

@frappe.whitelist()
def title1(args):
    date = datetime.strptime(args['start_date'],'%Y-%m-%d')
    date = date.strftime('%b-%Y').upper()
    if args['category']:
        data = ["SALARY SHEET OF "+ args['category'].upper()+" FOR THE MONTH OF " +date]
    else:
        data = ["SALARY SHEET OF EMPLOYEES FOR THE MONTH OF " +date]
    
    return data

@frappe.whitelist()
def get_data(args):
    data = []
    row = []
    row1=[]
    total_pd=0
    total_ot=0
    total_arrear_days=0
    total_arrear_overtime=0
    total_salary_ded=0
    total_shift_days=0
    total_basic_da=0
    total_da=0
    total_hra=0
    total_ca=0
    total_ma=0
    total_sa=0
    total_pb=0
    total_gross=0
    total_basic_da_ss=0
    total_da_ss=0
    total_ab_ss=0
    total_hra_ss=0
    total_ca_cc=0
    total_ma_ss=0
    total_sa_ss=0
    total_pb_ss=0
    total_inc=0
    total_ot_ss=0
    total_arrear_ot_ss=0
    total_arrear_salary=0
    total_shd=0
    total_gross_pay=0
    total_pf=0
    total_pt=0
    total_esi=0
    total_it_ded=0
    total_lwf=0
    total_salary_ded_ss=0
    total_ded=0
    total_ab=0
    total_cle=0
    total_ele=0
    total_bonus=0
    total_net_pay=0
    total_dhb=0


    start_date = datetime.strptime(args['start_date'], '%Y-%m-%d').date()
    salary_slip=frappe.db.get_all("Salary Slip",{'start_date':args['start_date'],'end_date':args['end_date']},['*'])
    indx=1
    for ss in salary_slip:
        if args['department']:
            emp = frappe.get_all("Employee",{'name':ss.employee,'department':args['department'],'status':'Active'},['*'],order_by='name ASC')
        elif args['category']:
            emp = frappe.get_all("Employee",{'name':ss.employee,'category':args['category'],'status':'Active'},['*'],order_by='name ASC')
        elif args['category'] and args['department']:
            emp = frappe.get_all("Employee",{'name':ss.employee,'department':args['department'],'category':args['category'],'status':'Active'},['*'],order_by='name ASC')	
        else:
            emp = frappe.get_all("Employee",{'name':ss.employee,'status':'Active'},['*'],order_by='name ASC')
        
        
        for e in emp:
            row=[indx,]
            row += [e.name,str(e.date_of_joining),e.employee_name]
            
            row.append(ss.payment_days or '')
            if ss.payment_days:
                total_pd+=ss.payment_days
            row.append(ss.overtime_hours or '-')
            if ss.overtime_hours:
                total_ot+=ss.overtime_hours
            row.append(ss.arrear_payment_days or '')
            if ss.arrear_payment_days:
                total_arrear_days+=ss.arrear_payment_days
            row.append(ss.arrear_overtime or '-')
            if ss.arrear_overtime:
                total_arrear_overtime+=ss.arrear_overtime
            row.append('-')
            row.append(ss.shift_days or '')
            if ss.shift_days:
                total_shift_days+=ss.shift_days
            basic_da = e.basic+e.dearness_allowance
            total_basic_da+=basic_da
            row.append(basic_da)
            row.append('-')
            row.append(e.house_rent_allowance or '-')
            if e.house_rent_allowance:
                total_hra+=e.house_rent_allowance
            row.append(e.conveyance_allowance or '-')
            if e.conveyance_allowance:
                total_ca+=e.conveyance_allowance
            row.append(e.medical_allowance or '-')
            if e.medical_allowance:
                total_ma+=e.medical_allowance
            row.append(e.special_allowance or '-')
            if e.special_allowance:
                total_sa+=e.special_allowance
            row.append(e.position_bonus or '-')
            if e.position_bonus:
                total_pb+=e.position_bonus
            row.append(e.gross)
            if e.gross:
                total_gross+=e.gross
            basic = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Basic"},['amount']) or 0
            da = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Dearness Allowance"},['amount']) or 0
            basic_da_ss=basic+da
            total_basic_da_ss+=basic_da_ss
            row.append(basic_da_ss)
            row.append("-")
            ab = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Arrear Basic"},['amount']) or 0
            if ab:
                total_ab_ss+=ab
            row.append(ab or '-')
            hra = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"House Rent Allowance"},['amount']) or 0
            if hra:
                total_hra_ss
            row.append(hra or '-')
            ca = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Conveyance Allowance"},['amount']) or 0
            if ca:
                total_ca_cc+=ca
            row.append(ca or '-')
            ma = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Medical Allowance"},['amount']) or 0
            if ma:
                total_ma_ss+=ma
            row.append(ma or '-')
            sa = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Special  Allowance"},['amount']) or 0
            if sa:
                total_sa_ss+=sa
            row.append(sa or '-')
            pb = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Position Bonus"},['amount']) or 0
            if pb:
                total_pb_ss+=pb
            row.append(pb or '-')
            inc = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Incentives"},['amount']) or 0
            if inc:
                total_inc+=inc
            row.append(inc or '-')
            ot = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Over Time"},['amount']) or 0
            if ot:
                total_ot_ss+=ot
            row.append(ot or '-')
            aot = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Arrear Overtime"},['amount']) or 0
            if aot:
                total_arrear_ot_ss+=aot
            row.append(aot or '-')
            arr = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Arrear"},['amount']) or 0
            if arr:
                total_arrear_salary+=arr
            row.append(arr or '-')
            sha = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Shift Allowance"},['amount']) or 0
            if sha:
                total_shd+=sha
            row.append(sha or '-')
            if ss.gross_pay:
                total_gross_pay+=ss.gross_pay
            row.append(ss.gross_pay)
            pf = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Provident Fund"},['amount']) or 0
            if pf:
                total_pf+=pf
            row.append(pf or '-')
            pt = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Professional Tax"},['amount']) or 0
            if pt:
                total_pt+=pt
            row.append(pt or '-')
            esi = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Employee State Insurance"},['amount']) or 0
            if esi:
                total_esi+=esi
            row.append(esi or '-')
            row.append("")
            lwf = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"LWF"},['amount']) or 0
            if lwf:
                total_lwf+=lwf
            row.append(lwf or '-')
            sad = frappe.db.get_value("Salary Detail",{'parent':ss.name,"salary_component":"Salary Advance Deduction"},['amount']) or 0
            if sad:
                total_salary_ded_ss+=sad
            row.append(sad or '-')
            if ss.total_deduction:
                total_ded+=ss.total_deduction
            row.append(ss.total_deduction)
            if ss.attendance_bonus:
                total_ab+=ss.attendance_bonus
            row.append(ss.attendance_bonus or '')
            if ss.casual_leave_encashment:
                total_cle+=ss.casual_leave_encashment
            row.append(ss.casual_leave_encashment or '')
            if ss.earned_leave_encashment:
                total_ele+=ss.earned_leave_encashment
            row.append(ss.earned_leave_encashment or '')
            if ss.bonus_yearly:
                total_bonus+=ss.bonus_yearly
            row.append(ss.bonus_yearly or '')
            if ss.net_pay:
                total_net_pay+=ss.net_pay
            row.append(ss.net_pay)
            if ss.declared_holiday_bonus:
                total_dhb+=ss.declared_holiday_bonus
            row.append(ss.declared_holiday_bonus or '')
            
            row.append(ss.department)
            
            row.append(e.payroll_cost_center)
            
            row.append(ss.department)
            data.append(row)
            indx +=1
    row1.append((indx-1))
    row1.append('')
    row1.append('')
    row1.append('')
    row1.append(total_pd)
    row1.append(total_ot)
    row1.append(total_arrear_days)
    row1.append(total_arrear_overtime)
    row1.append(total_salary_ded)
    row1.append(total_shift_days)
    row1.append(total_basic_da)
    row1.append(total_da)
    row1.append(total_hra)
    row1.append(total_ca)
    row1.append(total_ma)
    row1.append(total_sa)
    row1.append(total_pb)
    row1.append(total_gross)
    row1.append(total_basic_da_ss)
    row1.append(total_da_ss)
    row1.append(total_ab_ss)
    row1.append(total_hra_ss)
    row1.append(total_ca_cc)
    row1.append(total_ma_ss)
    row1.append(total_sa_ss)
    row1.append(total_pb_ss)
    row1.append(total_inc)
    row1.append(total_ot_ss)
    row1.append(total_arrear_ot_ss)
    row1.append(total_arrear_salary)
    row1.append(total_shd)
    row1.append(total_gross_pay)
    row1.append(total_pf)
    row1.append(total_pt)
    row1.append(total_esi)
    row1.append(total_it_ded)
    row1.append(total_lwf)
    row1.append(total_salary_ded_ss)
    row1.append(total_ded)
    row1.append(total_ab)
    row1.append(total_cle)
    row1.append(total_ele)
    row1.append(total_bonus)
    row1.append(total_net_pay)
    row1.append(total_dhb)
    data.append(row1)

    return data