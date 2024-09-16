import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Summary Of ESI Deduction Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    title_2=title2(args)
    ws.append(title_2)
    
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=6)

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data_1 = data1(args)
    for data in data_1:
        ws.append(data)
    ws.merge_cells(start_row=len(data_1)+3,start_column=1,end_row=len(data_1)+3,end_column=2)
    
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+3, min_col=1, max_col=6):
        for cell in rows:
            cell.border = border
    for rows in ws.iter_rows(min_row=len(data_1)+3,max_row=len(data_1)+3, min_col=1, max_col=2):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    for header in ws.iter_rows(min_row=len(data_1)+3,max_row=len(data_1)+3, min_col=1, max_col=6):
         for cell in header:
             cell.font = Font(bold=True)
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
    data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
    return data
@frappe.whitelist()
def title2(args):
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    mon = start_date.strftime('%B %Y').upper()
    data = ['SUMMARY OF ESI DEDUCTION FOR THE MONTH OF %s'%mon]
    return data
@frappe.whitelist()
def header1(args):
    data = ['SI.No.','Departments','Gross Salary',"EMPE'S ESI","EMPLR'S ESI",'TOTAL ESI']
    return data  
@frappe.whitelist()
def data1(args):
    data = []
    sa = frappe.get_all('Salary Slip', {'start_date':args.start_date,'end_date':args.end_date,'docstatus': ['!=', 2]}, ['salary_structure'])
    department_totals = {}
    
    for i in sa:
        ss= i.salary_structure
        structure_escaped = frappe.db.escape(ss)
        frappe.log_error(ss)

        def get_salary_component_total(salary_component):
            return frappe.db.sql(f"""
                select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
                left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
                where `tabSalary Slip`.salary_structure = {structure_escaped} and
                    `tabSalary Detail`.salary_component = '{salary_component}' and
                    `tabSalary Slip`.start_date >= '{args.start_date}' and
                    `tabSalary Slip`.end_date <= '{args.end_date}'
            """, as_dict=True)[0].total or 0
        gross = frappe.db.sql(f"""
            select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
            where `tabSalary Slip`.salary_structure = {structure_escaped} and
                `tabSalary Slip`.start_date >= '{args.start_date}' and
                `tabSalary Slip`.end_date <= '{args.end_date}'
        """, as_dict=True)[0].gross or 0
        frappe.log_error(gross)
        peesi = get_salary_component_total('Employee State Insurance')
        # if peesi == 0:
        #     plresi=0
        # else:
        #     plresi=
        plresi = get_salary_component_total('EMPLR Employee State Insurance')
        total_esi = peesi + plresi
        department_totals[ss] = [gross, peesi, plresi,total_esi]

    total_basic = sum(x[0] for x in department_totals.values())
    total_da = sum(x[1] for x in department_totals.values())
    total_arb = sum(x[2] for x in department_totals.values())
    total_hra = sum(x[3] for x in department_totals.values())
    
    for index, (ss, totals) in enumerate(department_totals.items(), start=1):
        data.append([index, ss]+ totals)

    row = [ 'Total','', total_basic, total_da, total_arb, total_hra,]
    data.append(row)
    return data