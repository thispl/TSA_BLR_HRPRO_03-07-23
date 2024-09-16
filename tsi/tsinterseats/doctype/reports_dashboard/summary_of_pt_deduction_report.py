import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Summary of PT Deduction Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['c'].width = 25

    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    title_2=title2(args)
    ws.append(title_2)
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=4):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data_1 = data1(args)
    for data in data_1:
        ws.append(data)
    bold_font = Font(bold=True)
    for cell in ws["11:11"]:
        cell.font = bold_font
    ws.merge_cells(start_row=11,start_column=1,end_row=11,end_column=2)

    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+2, min_col=1, max_col=4):
        for cell in rows:
            cell.border = border
    for rows in ws.iter_rows(min_row=len(data_1)+3,max_row=len(data_1)+3, min_col=3, max_col=4):
        for cell in rows:
            cell.border = border
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
    data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
    return data
@frappe.whitelist()
def title2(args):
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    mon = start_date.strftime('%B %Y').upper()
    data = ['Summary of PT deduction by no of employees for the month of %s'%mon]
    return data
@frappe.whitelist()
def header1(args):
    data = ['Professional Tax Details','No of Employee','Amount','Total']
    return data  
@frappe.whitelist()
def data1(args):
    data = []
    sa = frappe.get_all('Salary Slip', {'start_date': args.start_date, 'end_date': args.end_date, 'docstatus': ['!=', 2]}, ['*'])
    department_totals = {}
    salary_structure_count = {}
    index = 1 
    if not isinstance(sa, list):
        sa = [sa]
    
    for i in sa:
        ss = i.name
        if ss in salary_structure_count:
            salary_structure_count[ss] += 1
        else:
            salary_structure_count[ss] = 1
    total_count = sum(salary_structure_count.values())
    pt = frappe.db.sql("""select `tabSalary Detail`.amount as pt from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Detail`.salary_component = 'Professional Tax' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(args.start_date,args.end_date),as_dict=True)[0].pt or 0
    total=total_count*pt
    data.append(['1. Less than Rs.3000', 0,0,0])
    data.append(['2. Not less than Rs.3000 but less than Rs.5000', 0,0, 0])
    data.append(['3. Not less than Rs.5000 but less than Rs.8000', 0,0, 0])
    data.append(['4. Not less than Rs.8000 but less than Rs.10000', 0,0, 0])
    data.append(['1. Upto 24999- NIL', 0,0, 0])
    if pt == 200:
        row=(['2. Rs. 25000 and above- Rs. 200 per month', total_count,pt,total or 0])

    data.append(row)
    data.append(['3. Others(THAI TEAM)', 0,0, 0])

    data.append(['','','Net amount payable',total])
    return data