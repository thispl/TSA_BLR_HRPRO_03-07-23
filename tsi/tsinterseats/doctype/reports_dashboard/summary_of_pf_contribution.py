from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download():
	filename = 'Summary of PF Contribution'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['B'].width = 18 
	ws.column_dimensions['C'].width = 16 
	ws.column_dimensions['D'].width = 16
	ws.column_dimensions['E'].width = 24
	ws.column_dimensions['F'].width = 15
	ws.column_dimensions['G'].width = 24
	ws.column_dimensions['H'].width = 15
	ws.column_dimensions['I'].width = 15
	ws.column_dimensions['J'].width = 20
	header_date = title (args)
	ws.append(header_date) 

	ws.append(title1(args))

	ws.append(header1(args))
	ws.append(['','','','','1','2','3','4','5','6'])
	data_1 = data1(args)
	for data in data_1:
		ws.append(data)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10 )
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(data_1)+4, max_row=len(data_1)+4, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=4, max_row=len(data_1)+4, min_col=1, max_col=10):
		for cell in header:
			cell.alignment = align_center
	border = Border(left=Side(border_style='thin', color='000000'),
			right=Side(border_style='thin', color='000000'),
			top=Side(border_style='thin', color='000000'),
			bottom=Side(border_style='thin', color='000000')) 
	for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+4, min_col=1, max_col=10):
		for cell in rows:
			cell.border = border
	
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title(args):
	data = ["TS Interseats India Private Limited"]
	return data

@frappe.whitelist()
def title1(args):
	date = datetime.strptime(args['start_date'],'%Y-%m-%d')
	date = date.strftime('%B %Y')
	data = ["Summary of PF Contribution for the Month of  " +date]
	
	return data

@frappe.whitelist()
def header1(args):
	data = ['S.No.','Department','No of persons','BASIC SAL.','Employee PF 12%','VPF','Employee PF (Incl.PF)','Family PF','EPF @ 3.67','"A/c. 10 Basic Salary"']
	return data

@frappe.whitelist()
def data1(args):
	data = []

	# sa = frappe.get_all('Salary Slip', ['department'])
	sa = frappe.db.sql("""select * from `tabSalary Slip` where start_date between '%s' and '%s' and docstatus!=2 """%(args.start_date, args.end_date), as_dict=True)
	ab=0
	department_totals = {}
	employee_count={}
	for i in sa:
		ss = i.payroll_category
		department_escaped = frappe.db.escape(ss)

		def get_salary_component_total(salary_component):
			return frappe.db.sql(f"""
				select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
				left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
				where 
				`tabSalary Slip`.Payroll_category = {department_escaped} and
				`tabSalary Detail`.salary_component = '{salary_component}' and 
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
			""", as_dict=True)[0].total or 0

		b = get_salary_component_total('Basic')
		hra = get_salary_component_total('House Rent Allowance')
		ot = get_salary_component_total('Overtime')
		aot = get_salary_component_total('Arrear Overtime')
		gross = frappe.db.sql(f"""
			select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
			where `tabSalary Slip`.payroll_category = {department_escaped} and
				`tabSalary Slip`.start_date >= '{args.start_date}' and
				`tabSalary Slip`.end_date <= '{args.end_date}'
		""", as_dict=True)[0].gross or 0

		val=(gross-ot-aot-hra)*0.12
		if val==0:
			pf=1800
		tt=(gross-ot-aot-hra)
		if tt>15000:
			ab=15000
		pt = get_salary_component_total('VPF')
		esi=ab*0.12
		lwf=ab*0.0833
		sa=esi-lwf
		lwf = get_salary_component_total('LWF')
		sd = get_salary_component_total('salary Deduction')
		
		if ss in employee_count:
			employee_count[ss] += 1
		else:
			employee_count[ss] = 1		

		department_totals[ss] = [b,  pf, pt, esi, lwf, sd,  ab,]

	# Calculate the grand totals for all departments
	total_basic = sum(x[0] for x in department_totals.values())
	total_da = sum(x[1] for x in department_totals.values())
	total_arb = sum(x[2] for x in department_totals.values())
	total_sa = sum(x[3] for x in department_totals.values())
	total_i = sum(x[4] for x in department_totals.values())
	total_a = sum(x[5] for x in department_totals.values())
	total_ot = sum(x[6] for x in department_totals.values())
	for index, (ss, totals) in enumerate(department_totals.items(), start=1):
		data.append([index, ss, employee_count[ss]] + totals)

	employee_counts = [employee_count[ss] for ss in department_totals.keys()]
	total_counts = sum(employee_counts)
	row = ['', 'Grand Total',total_counts , total_basic, total_da, total_arb, total_sa, total_i, total_a, total_ot, ]
	
	data.append(row)

	return data