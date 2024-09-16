import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Summary of Labour Welfare Fund deduction Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['B'].width = 35
    title_2=title2(args)
    ws.append(title_2)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    header_1=header1(args)
    ws.append(header_1)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=2)

    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data_1 = data1(args)
    for data in data_1:
        ws.append(data)
    for rows in ws.iter_rows(min_row=3,max_row=len(data_1), min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    for rows in ws.iter_rows(min_row=1, max_row=len(data_1)+3, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=len(data_1)+2, start_column=1, end_row=len(data_1)+2, end_column=2)
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for rows in ws.iter_rows(min_row=1,max_row=len(data_1)+2, min_col=1, max_col=6):
        for cell in rows:
            cell.border = border
    for header in ws.iter_rows(min_row=len(data_1)+2, max_row=len(data_1)+2, min_col=1, max_col=6):
         for cell in header:
             cell.font = Font(bold=True)
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'



@frappe.whitelist()
def title2(args):
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    mon = start_date.strftime('%Y').upper()
    data = ['Summary of Labour Welfare Fund deduction by no of employees for the year %s'%mon]
    return data
@frappe.whitelist()
def header1(args):
    data = ['Labour Welfare Fund Contribution details','','No.Of Employees','Employee Contribution @ Rs.20/-','Employer Contribution @ Rs.40/-','Total']
    return data
@frappe.whitelist()
def data1(args):
    data = []
    sa = frappe.get_all('Salary Slip', {'start_date': args.start_date, 'end_date': args.end_date, 'docstatus': ['!=', 2]}, ['salary_structure'])
    department_totals = {}
    salary_structure_count = {}
    index = 1 
    if not isinstance(sa, list):
        sa = [sa]

    for i in sa:
        ss = i.salary_structure
        if ss in salary_structure_count:
            salary_structure_count[ss] += 1
        else:
            salary_structure_count[ss] = 1

    for ss, count in salary_structure_count.items():
        multiplied_count_20 = count * 20
        multiplied_count_40 = count * 40
        total = multiplied_count_20 + multiplied_count_40
        data.append([index, ss, count, multiplied_count_20, multiplied_count_40, total])
        index += 1 

    salary_structure_counts = [count for count in salary_structure_count.values()]
    e_count=sum(salary_structure_counts)
    total_counts_20 = sum(salary_structure_counts) * 20
    total_counts_40 = sum(salary_structure_counts) * 40
    total_count = total_counts_20 + total_counts_40
    data.append(['Total Contribution payable','', e_count,total_counts_20 , total_counts_40, total_count])

    return data