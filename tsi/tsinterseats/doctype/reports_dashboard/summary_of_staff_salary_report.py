import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold,msgprint
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Summary Of Staff Salary Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=29)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=29):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    title_2=title2(args)
    ws.append(title_2)
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=29)
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=29):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=29):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data_1 = data1(args)
    for data in data_1:
        ws.append(data)
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for rows in ws.iter_rows(min_row=2,max_row=len(data_1)+3, min_col=1, max_col=29):
        for cell in rows:
            cell.border = border
    for header in ws.iter_rows(min_row=len(data_1)+3, max_row=len(data_1)+3, min_col=1, max_col=29):
         for cell in header:
             cell.font = Font(bold=True)
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
    data = ['TS INTERSEATS INDIA PRIVATE LIMITED']
    return data
@frappe.whitelist()
def title2(args):
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    mon = start_date.strftime('%B %Y').upper()
    data = ['SUMMARY OF STAFF SALARY FOR THE MONTH OF %s'%mon]
    return data
@frappe.whitelist()
def header1(args):
    data = ['Sr.No.','Departments','No.Of Employees','Basic Salary','Dearness Allowance','Arr.Basic Salary & DA','House Rent Allowance','Conveyance Allowance','Medical Allowance','Special Allowance','Position Bonus','Incentives','Over Time Salary','Arr. Over Time Salary','Arrears Salary','Shift Allowance','Gross Salary','PF','PT','ESI','TDS Deduction','LWF','Salary Deduction','Deductions','Attendance Bonus','Casual Leave Encashment','Earned Leave Encashment','Bonus Yearly','Net Salary']
    return data
# @frappe.whitelist()
# def data1(args):
# 	data = []
# 	sa = frappe.get_all('Salary Slip', {'start_date':args.start_date,'end_date':args.end_date}, ['salary_structure'])
# 	department_totals = {}
    
# 	for i in sa:
# 		ss= i.salary_structure
# 		structure_escaped = frappe.db.escape(ss)

# 		def get_salary_component_total(salary_component):
# 			return frappe.db.sql(f"""
# 				select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
# 				left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
# 				where `tabSalary Slip`.salary_structure = {structure_escaped} and
# 					`tabSalary Detail`.salary_component = '{salary_component}' and
# 					`tabSalary Slip`.start_date >= '{args.start_date}' and
# 					`tabSalary Slip`.end_date <= '{args.end_date}'
# 			""", as_dict=True)[0].total or 0
# 		b = get_salary_component_total('Basic')
# 		da = get_salary_component_total('Dearness Allowance')
        
# 		arb = get_salary_component_total('Arrear Basic')
# 		hra = get_salary_component_total('House Rent Allowance')
# 		ca = get_salary_component_total('conveyance Allowance')
# 		ma = get_salary_component_total('Medical Allowance')
# 		sa = get_salary_component_total('Special Allowance')
# 		pb = get_salary_component_total('Position Bonus')

# 		i = get_salary_component_total('Incentives')
# 		ot = get_salary_component_total('Overtime')
       
# 		aot = get_salary_component_total('Arrear Overtime')
# 		a = get_salary_component_total('Arrear')

# 		sha = get_salary_component_total('Shift Allowance')
# 		gross = frappe.db.sql(f"""
# 			select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
# 			where `tabSalary Slip`.department = {structure_escaped} and
# 				`tabSalary Slip`.start_date >= '{args.start_date}' and
# 				`tabSalary Slip`.end_date <= '{args.end_date}'
# 		""", as_dict=True)[0].gross or 0

# 		pf = get_salary_component_total('Provident Fund')
# 		pt = get_salary_component_total('Professional Tax')
# 		esi = get_salary_component_total('Employee State Insurance')
# 		tds = get_salary_component_total('TDS')

# 		lwf = get_salary_component_total('LWF')
# 		sd = get_salary_component_total('Salary Deduction')

# 		deduction = frappe.db.sql(f"""
# 			select sum(`tabSalary Slip`.total_deduction) as deduction from `tabSalary Slip`
# 			where `tabSalary Slip`.department = {structure_escaped} and
# 				`tabSalary Slip`.start_date >= '{args.start_date}' and
# 				`tabSalary Slip`.end_date <= '{args.end_date}'
# 		""", as_dict=True)[0].deduction or 0

# 		ab = get_salary_component_total('Attendance Bonus')
# 		cle = get_salary_component_total('Causal Leave Encashment')
# 		ele = get_salary_component_total('Earned Leave Encashment')
# 		b = get_salary_component_total(' Bonus')
# 		npay = frappe.db.sql(f"""
# 			select sum(`tabSalary Slip`.net_pay) as npay from `tabSalary Slip`
# 			where `tabSalary Slip`.department = {structure_escaped} and
# 				`tabSalary Slip`.start_date >= '{args.start_date}' and
# 				`tabSalary Slip`.end_date <= '{args.end_date}'
# 		""", as_dict=True)[0].npay or 0
# 		department_totals[ss] = [b, da, arb,hra,ca,ma, sa,pb, i,  ot, aot,a, sha, gross, pf, pt, esi,tds, lwf,sd, deduction, ab, cle, ele,b,npay,]

# 	total_basic = sum(x[0] for x in department_totals.values())
# 	total_da = sum(x[1] for x in department_totals.values())
# 	total_arb = sum(x[2] for x in department_totals.values())
# 	total_hra = sum(x[3] for x in department_totals.values())
# 	total_ca = sum(x[4] for x in department_totals.values())
# 	total_ma = sum(x[6] for x in department_totals.values())
# 	total_sa = sum(x[5] for x in department_totals.values())
# 	total_pb = sum(x[7] for x in department_totals.values())
# 	total_i = sum(x[8] for x in department_totals.values())
# 	total_ot = sum(x[9] for x in department_totals.values())
# 	total_aot = sum(x[10] for x in department_totals.values())
# 	total_a = sum(x[11] for x in department_totals.values())
# 	total_sha = sum(x[12] for x in department_totals.values())
# 	total_gross = sum(x[13] for x in department_totals.values())
# 	total_pf = sum(x[14] for x in department_totals.values())
# 	total_pt = sum(x[15] for x in department_totals.values())
# 	total_esi = sum(x[16] for x in department_totals.values())
# 	total_lwf = sum(x[17] for x in department_totals.values())
# 	total_sd = sum(x[18] for x in department_totals.values())
# 	total_deduction = sum(x[19] for x in department_totals.values())
# 	total_ab = sum(x[20] for x in department_totals.values())
# 	total_cle = sum(x[21] for x in department_totals.values())
# 	total_ele = sum(x[22] for x in department_totals.values())
# 	total_b = sum(x[23] for x in department_totals.values())
# 	total_npay = sum(x[24] for x in department_totals.values())
# 	for index, (ss, totals) in enumerate(department_totals.items(), start=1):
# 		data.append([index, ss]+ totals)

# 		row = ['', 'Grand Total', total_basic, total_da, total_arb, total_hra, total_ca, total_ma, total_sa, total_pb, total_i, total_ot, total_aot, total_a, total_sha, total_gross, total_pf, total_pt, total_esi, total_lwf, total_sd, total_deduction, total_ab, total_cle, total_ele, total_b, total_npay]
# 	data.append(row)
# 	return data

@frappe.whitelist()
def data1(args):
    data = []

    sa = frappe.get_all('Salary Slip', {'salary_structure': "Staff",'start_date': args.start_date, 'end_date': args.end_date}, ['salary_structure'])
    
    department_totals = {}
    salary_structure_count = {}

    if not isinstance(sa, list):
        sa = [sa]

    for i in sa:
        ss = i.salary_structure
        structure_escaped = frappe.db.escape(ss)

        def get_salary_component_total(salary_component):
            return frappe.db.sql(f"""
                select sum(`tabSalary Detail`.amount) as total from `tabSalary Slip`
                left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent
                where `tabSalary Slip`.salary_structure = {structure_escaped} and
                    `tabSalary Detail`.salary_component = '{salary_component}' and
                    `tabSalary Slip`.start_date >= '{args.start_date}' and
                    `tabSalary Slip`.end_date <= '{args.end_date}'
            """, as_dict=True)[0].total or 0
        
        b = get_salary_component_total('Basic')
        da = get_salary_component_total('Dearness Allowance')
        arb = get_salary_component_total('Arrear Basic')
        hra = get_salary_component_total('House Rent Allowance')
        ca = get_salary_component_total('conveyance Allowance')
        ma = get_salary_component_total('Medical Allowance')
        sa = get_salary_component_total('Special  Allowance')
        pb = get_salary_component_total('Position Bonus')
        i = get_salary_component_total('Incentives')
        ot = get_salary_component_total('Overtime')
        aot = get_salary_component_total('Arrear Overtime')
        a = get_salary_component_total('Arrear')
        sha = get_salary_component_total('Shift Allowance')

        gross = frappe.db.sql(f"""
            select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
            where `tabSalary Slip`.salary_structure = {structure_escaped} and
                `tabSalary Slip`.start_date >= '{args.start_date}' and
                `tabSalary Slip`.end_date <= '{args.end_date}'
        """, as_dict=True)[0].gross or 0

        pf = get_salary_component_total('Provident Fund')
        pt = get_salary_component_total('Professional Tax')
        esi = get_salary_component_total('Employee State Insurance')
        tds = get_salary_component_total('TDS')
        lwf = get_salary_component_total('LWF')
        sd = get_salary_component_total('Salary Deduction')

        deduction = frappe.db.sql(f"""
            select sum(`tabSalary Slip`.total_deduction) as deduction from `tabSalary Slip`
            where `tabSalary Slip`.salary_structure = {structure_escaped} and
                `tabSalary Slip`.start_date >= '{args.start_date}' and
                `tabSalary Slip`.end_date <= '{args.end_date}'
        """, as_dict=True)[0].deduction or 0

        ab = get_salary_component_total('Attendance Bonus')
        cle = get_salary_component_total('Causal Leave Encashment')
        ele = get_salary_component_total('Earned Leave Encashment')
        bo = get_salary_component_total(' Bonus')
        
        npay = frappe.db.sql(f"""
            select sum(`tabSalary Slip`.net_pay) as npay from `tabSalary Slip`
            where `tabSalary Slip`.salary_structure = {structure_escaped} and
                `tabSalary Slip`.start_date >= '{args.start_date}' and
                `tabSalary Slip`.end_date <= '{args.end_date}'
        """, as_dict=True)[0].npay or 0
        
        if ss in salary_structure_count:
            salary_structure_count[ss] += 1
        else:
            salary_structure_count[ss] = 1

        department_totals[ss] = [b, da, arb, hra, ca, ma, sa, pb, i, ot, aot, a, sha, gross, pf, pt, esi, tds, lwf, sd, deduction, ab, cle, ele, bo, npay]

    total_basic = sum(x[0] for x in department_totals.values())
    total_da = sum(x[1] for x in department_totals.values())
    total_arb = sum(x[2] for x in department_totals.values())
    total_hra = sum(x[3] for x in department_totals.values())
    total_ca = sum(x[4] for x in department_totals.values())
    total_ma = sum(x[5] for x in department_totals.values())
    total_sa = sum(x[6] for x in department_totals.values())
    total_pb = sum(x[7] for x in department_totals.values())
    total_i = sum(x[8] for x in department_totals.values())
    total_ot = sum(x[9] for x in department_totals.values())
    total_aot = sum(x[10] for x in department_totals.values())
    total_a = sum(x[11] for x in department_totals.values())
    total_sha = sum(x[12] for x in department_totals.values())
    total_gross = sum(x[13] for x in department_totals.values())
    total_pf = sum(x[14] for x in department_totals.values())
    total_pt = sum(x[15] for x in department_totals.values())
    total_esi = sum(x[16] for x in department_totals.values())
    total_tds = sum(x[17] for x in department_totals.values())

    total_lwf = sum(x[18] for x in department_totals.values())
    total_sd = sum(x[19] for x in department_totals.values())
    total_deduction = sum(x[20] for x in department_totals.values())
    total_ab = sum(x[21] for x in department_totals.values())
    total_cle = sum(x[22] for x in department_totals.values())
    total_ele = sum(x[23] for x in department_totals.values())
    total_bo = sum(x[24] for x in department_totals.values())
    total_npay = sum(x[25] for x in department_totals.values())

    for index, (ss, totals) in enumerate(department_totals.items(), start=1):
        data.append([index, ss, salary_structure_count[ss]] + totals)

    salary_structure_counts = [salary_structure_count[ss] for ss in department_totals.keys()]
    total_counts = sum(salary_structure_counts)
    grand_total_row = ['', 'Grand Total'] + [total_counts]  + [
        total_basic, total_da, total_arb, total_hra, total_ca, total_ma, total_sa, total_pb, total_i, total_ot,
        total_aot, total_a, total_sha, total_gross, total_pf, total_pt, total_esi,total_tds ,total_lwf, total_sd,
        total_deduction, total_ab, total_cle, total_ele, total_bo, total_npay
    ]
    data.append(grand_total_row)

    return data